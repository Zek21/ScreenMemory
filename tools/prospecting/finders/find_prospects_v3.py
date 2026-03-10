"""
UAE Business Finder v3 - FAST list-level extraction.
Checks for website presence directly in the search results feed,
only clicks into businesses WITHOUT a visible website link for details.
10x faster than clicking every result.
"""

import asyncio
import json
import os
import re
import sys
import csv
from datetime import datetime
from urllib.parse import quote_plus
from playwright.async_api import async_playwright, TimeoutError as PwTimeout

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
CATEGORIES_FILE = os.path.join(SCRIPT_DIR, "categories.txt")
OUTPUT_DIR = os.path.join(SCRIPT_DIR, "results", "uae", "_intermediate")
CITIES = ["Dubai", "Abu Dhabi", "Sharjah", "Ajman", "Ras Al Khaimah", "Fujairah", "Umm Al Quwain"]
MAX_SCROLL = 4
BATCH_SIZE = 3

TEST_MODE = "--test" in sys.argv
if TEST_MODE:
    CITIES = ["Dubai"]


def load_categories(path):
    with open(path, "r", encoding="utf-8") as f:
        cats = [line.strip() for line in f if line.strip()]
    return cats[:5] if TEST_MODE else cats


def progress_path():
    return os.path.join(OUTPUT_DIR, "progress.json")


def load_progress():
    p = progress_path()
    if os.path.exists(p):
        with open(p) as f:
            return json.load(f)
    return {"completed": []}


def save_progress(data):
    with open(progress_path(), "w") as f:
        json.dump(data, f, indent=2)


async def safe_goto(page, url, retries=2):
    for attempt in range(retries):
        try:
            await page.goto(url, wait_until="domcontentloaded", timeout=20000)
            return True
        except Exception:
            await asyncio.sleep(2)
    return False


async def search_and_extract(page, category, city, seen):
    """Search Google Maps - extract businesses WITHOUT website from list + click for details."""
    results = []
    query = f"{category} in {city} UAE"
    url = f"https://www.google.com/maps/search/{quote_plus(query)}"

    if not await safe_goto(page, url):
        return results

    await asyncio.sleep(3)

    # Dismiss consent
    try:
        consent = page.locator('button:has-text("Accept all")')
        if await consent.count() > 0:
            await consent.first.click(timeout=3000)
            await asyncio.sleep(1)
    except Exception:
        pass

    # Wait for feed
    feed = page.locator('div[role="feed"]')
    try:
        await feed.wait_for(state="attached", timeout=6000)
    except PwTimeout:
        return results

    # Scroll to load results
    for _ in range(MAX_SCROLL):
        try:
            await feed.evaluate("el => el.scrollTop = el.scrollHeight")
            await asyncio.sleep(1)
        except Exception:
            break

    # FAST: Extract ALL business data from the list using JavaScript
    # This checks each article in the feed for website link presence
    list_data = await page.evaluate(r"""() => {
        const articles = document.querySelectorAll('div[role="feed"] > div:not(:first-child)');
        const businesses = [];
        
        for (const article of articles) {
            // Skip sponsored
            if (article.querySelector('h1') && article.innerText.includes('Sponsored')) continue;
            
            // Get the place link
            const placeLink = article.querySelector('a[href*="/maps/place/"]');
            if (!placeLink) continue;
            
            const name = (placeLink.getAttribute('aria-label') || '').trim();
            if (!name) continue;
            
            // CHECK: does this list entry have a "Website" link?
            const websiteLink = article.querySelector('a[href]:not([href*="google.com"])');
            const websiteBtn = article.querySelector('a[aria-label*="website" i], a[aria-label*="Visit" i]');
            const websiteText = article.querySelector('a');
            let hasWebsite = false;
            
            // Look for "Website" text in buttons/links within this article
            const allLinks = article.querySelectorAll('a');
            for (const link of allLinks) {
                const text = link.innerText.trim().toLowerCase();
                const label = (link.getAttribute('aria-label') || '').toLowerCase();
                if (text === 'website' || label.includes('website') || label.includes("visit")) {
                    hasWebsite = true;
                    break;
                }
            }
            
            // Extract basic info from list
            const text = article.innerText;
            let phone = '';
            const phoneMatch = text.match(/(\+?\d{1,4}[\s\-]?\d{2,4}[\s\-]?\d{3,4}[\s\-]?\d{3,4})/);
            if (phoneMatch) phone = phoneMatch[1].trim();
            
            let rating = '';
            const ratingMatch = text.match(/^(\d\.\d)/m);
            if (ratingMatch) rating = ratingMatch[1];
            
            let reviews = '';
            const reviewMatch = text.match(/\(([\d,]+)\)/);
            if (reviewMatch) reviews = reviewMatch[1];
            
            // Get category/type from list text (usually appears as a line)
            let gType = '';
            const lines = text.split('\n').map(l => l.trim()).filter(l => l);
            for (const line of lines) {
                if (line.includes('·') && !line.includes('star') && !line.includes('Open') && !line.includes('Close')) {
                    const parts = line.split('·').map(p => p.trim());
                    if (parts[0] && parts[0].length < 40 && !parts[0].match(/^\d/)) {
                        gType = parts[0];
                        break;
                    }
                }
            }
            
            // Extract address snippet
            let address = '';
            for (const line of lines) {
                if (line.includes('·') && (line.includes('St') || line.includes('Rd') || line.includes('area') || line.includes('near'))) {
                    const parts = line.split('·');
                    address = parts[parts.length - 1].trim();
                    break;
                }
            }
            
            businesses.push({
                name, hasWebsite, phone, rating, reviews, gType, address,
                linkIndex: Array.from(document.querySelectorAll('div[role="feed"] a[href*="/maps/place/"]')).indexOf(placeLink)
            });
        }
        return businesses;
    }""")

    no_website = [b for b in list_data if not b.get("hasWebsite") and b["name"] not in seen]
    with_website = [b for b in list_data if b.get("hasWebsite")]
    
    print(f"({len(list_data)} total, {len(no_website)} no-web, {len(with_website)} has-web)", end=" ")

    # For businesses WITHOUT website in list, click for full details
    for biz in no_website:
        name = biz["name"]
        if name in seen:
            continue

        try:
            # Click the link by index
            links = page.locator('div[role="feed"] a[href*="/maps/place/"]')
            count = await links.count()
            
            # Find the link with matching aria-label
            clicked = False
            for li in range(count):
                label = await links.nth(li).get_attribute("aria-label") or ""
                if label.strip() == name:
                    await links.nth(li).click(timeout=5000)
                    clicked = True
                    break
            
            if not clicked:
                continue
                
            await asyncio.sleep(2)

            # Wait for detail panel
            try:
                await page.wait_for_function(
                    """() => {
                        const h1s = document.querySelectorAll('h1');
                        for (const h1 of h1s) {
                            const r = h1.getBoundingClientRect();
                            if (r.left > 400 && h1.innerText.trim().length > 1 && h1.innerText.trim() !== 'Sponsored') return true;
                        }
                        return false;
                    }""",
                    timeout=5000
                )
            except Exception:
                pass

            # Double-check website in detail panel + get full address
            detail = await page.evaluate(r"""() => {
                const result = {hasWebsite: false, address: '', phone: '', fullName: ''};
                
                // Full name from detail panel
                const h1s = document.querySelectorAll('h1');
                for (const h1 of h1s) {
                    const r = h1.getBoundingClientRect();
                    if (r.left > 400 && h1.innerText.trim().length > 1 && h1.innerText.trim() !== 'Sponsored') {
                        result.fullName = h1.innerText.trim();
                        break;
                    }
                }
                
                // Website check in detail panel
                const webLink = document.querySelector('a[data-item-id="authority"]');
                const webBtn = document.querySelector('button[data-item-id="authority"]');
                result.hasWebsite = !!(webLink || webBtn);
                
                // Full address
                const addrBtn = document.querySelector('button[data-item-id="address"]');
                if (addrBtn) {
                    result.address = (addrBtn.getAttribute('aria-label') || '').replace(/Address:\s*/i, '').trim();
                }
                
                // Phone (more reliable from detail)
                const phoneBtn = document.querySelector('button[data-item-id^="phone"]');
                if (phoneBtn) {
                    result.phone = (phoneBtn.getAttribute('aria-label') || '').replace(/Phone:\s*/i, '').trim();
                }
                
                return result;
            }""")

            # Only add if CONFIRMED no website in detail view
            if detail and not detail.get("hasWebsite"):
                final_name = detail.get("fullName") or name
                seen.add(final_name)
                seen.add(name)
                results.append({
                    "name": final_name,
                    "search_category": category,
                    "google_category": biz.get("gType", ""),
                    "city": city,
                    "address": detail.get("address") or biz.get("address", ""),
                    "phone": detail.get("phone") or biz.get("phone", ""),
                    "rating": biz.get("rating", ""),
                    "reviews": biz.get("reviews", ""),
                    "has_website": "No",
                    "found_date": datetime.now().strftime("%Y-%m-%d"),
                })
            else:
                seen.add(name)

            # Go back to results
            back = page.locator('button[aria-label="Back"]')
            try:
                if await back.count() > 0:
                    await back.first.click(timeout=3000)
                    await asyncio.sleep(1.5)
                else:
                    await page.go_back()
                    await asyncio.sleep(2)
                # Verify feed is back
                try:
                    await page.locator('div[role="feed"]').wait_for(state="attached", timeout=3000)
                except PwTimeout:
                    await safe_goto(page, url)
                    await asyncio.sleep(2)
            except Exception:
                await safe_goto(page, url)
                await asyncio.sleep(2)

        except Exception as e:
            seen.add(name)
            try:
                await safe_goto(page, url)
                await asyncio.sleep(2)
            except Exception:
                break

    return results


def append_csv(results, path):
    if not results:
        return
    fields = ["name", "search_category", "google_category", "city", "address",
              "phone", "rating", "reviews", "has_website", "found_date"]
    exists = os.path.exists(path)
    with open(path, "a", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        if not exists:
            w.writeheader()
        w.writerows(results)


def csv_to_excel(csv_path, xlsx_path):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "UAE Prospects"

        headers = ["Business Name", "Search Category", "Google Category", "City",
                   "Address", "Phone", "Rating", "Reviews", "Has Website", "Date Found"]
        hfill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        hfont = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))

        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.fill, cell.font, cell.alignment, cell.border = hfill, hfont, Alignment(horizontal="center"), border

        if os.path.exists(csv_path):
            with open(csv_path, "r", encoding="utf-8") as f:
                keys = ["name", "search_category", "google_category", "city",
                        "address", "phone", "rating", "reviews", "has_website", "found_date"]
                for ri, row in enumerate(csv.DictReader(f), 2):
                    for ci, k in enumerate(keys, 1):
                        cell = ws.cell(row=ri, column=ci, value=row.get(k, ""))
                        cell.border = border
                        cell.font = Font(name="Calibri", size=10)

        for i, w in enumerate([35, 22, 22, 16, 45, 18, 8, 10, 12, 12], 1):
            ws.column_dimensions[chr(64+i)].width = w
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        wb.save(xlsx_path)
        print(f"\n  Excel: {xlsx_path}")
    except Exception as e:
        print(f"\n  Excel error: {e}")


def count_rows(path):
    if not os.path.exists(path):
        return 0
    with open(path, encoding="utf-8") as f:
        return max(0, sum(1 for _ in f) - 1)


async def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    categories = load_categories(CATEGORIES_FILE)
    progress = load_progress()
    done = set(progress.get("completed", []))
    remaining = [c for c in categories if c not in done]

    csv_path = os.path.join(OUTPUT_DIR, "uae_businesses_no_website.csv")
    xlsx_path = os.path.join(OUTPUT_DIR, "uae_businesses_no_website.xlsx")

    mode = "TEST" if TEST_MODE else "FULL"
    print(f"\n{'='*60}")
    print(f"  UAE BUSINESS FINDER v3 - {mode}")
    print(f"  {len(remaining)} categories | Cities: {', '.join(CITIES)}")
    print(f"{'='*60}\n")

    if not remaining:
        print("All done! Regenerating Excel...")
        csv_to_excel(csv_path, xlsx_path)
        return

    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=False,
            args=["--disable-blink-features=AutomationControlled", "--lang=en-US",
                  "--no-first-run", "--disable-extensions"]
        )
        ctx = await browser.new_context(
            viewport={"width": 1366, "height": 900},
            locale="en-US",
            timezone_id="Asia/Dubai",
            geolocation={"latitude": 25.2048, "longitude": 55.2708},
            permissions=["geolocation"],
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        )
        await ctx.add_init_script("Object.defineProperty(navigator,'webdriver',{get:()=>undefined})")
        page = await ctx.new_page()
        page.set_default_timeout(10000)

        await safe_goto(page, "https://www.google.com/maps")
        await asyncio.sleep(2)

        seen = set()
        batch_results = []
        batch_n = 0

        for idx, cat in enumerate(remaining):
            print(f"\n[{idx+1}/{len(remaining)}] {cat.upper()}")
            for city in CITIES:
                print(f"  {city}: ", end="", flush=True)
                try:
                    found = await search_and_extract(page, cat, city, seen)
                    print(f"-> {len(found)} prospects")
                    batch_results.extend(found)
                except Exception as e:
                    print(f"FAIL: {e}")

            done.add(cat)
            batch_n += 1

            if batch_n >= BATCH_SIZE or idx == len(remaining) - 1:
                if batch_results:
                    append_csv(batch_results, csv_path)
                    total = count_rows(csv_path)
                    print(f"  >> SAVED {len(batch_results)} new | {total} total")
                    batch_results = []
                batch_n = 0
                progress["completed"] = list(done)
                save_progress(progress)

        csv_to_excel(csv_path, xlsx_path)
        total = count_rows(csv_path)
        print(f"\n{'='*60}")
        print(f"  DONE! {total} businesses WITHOUT websites found")
        print(f"  CSV:   {csv_path}")
        print(f"  Excel: {xlsx_path}")
        print(f"{'='*60}")
        await browser.close()


if __name__ == "__main__":
    asyncio.run(main())
